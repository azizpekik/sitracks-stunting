import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from typing import List, Dict, Any
from datetime import datetime
import logging

from models import Measurement, Child

logger = logging.getLogger(__name__)


class ReportGenerator:
    def __init__(self):
        # Define color schemes for validation status
        self.colors = {
            'ERROR': PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid'),  # Red
            'WARNING': PatternFill(start_color='FFEB3B', end_color='FFEB3B', fill_type='solid'),  # Yellow
            'WARNING_ANOMALY': PatternFill(start_color='FF9800', end_color='FF9800', fill_type='solid'),  # Orange
            'OK': PatternFill(start_color='4CAF50', end_color='4CAF50', fill_type='solid'),  # Green
            'MISSING': PatternFill(start_color='FFEB3B', end_color='FFEB3B', fill_type='solid'),  # Yellow
        }

        # Define border style
        self.thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Define header font
        self.header_font = Font(bold=True, color='FFFFFF')
        self.header_fill = PatternFill(start_color='2196F3', end_color='2196F3', fill_type='solid')  # Blue

    async def generate_excel_report(self, file_path: str, measurements: List[Measurement], default_gender: str):
        """
        Generate Excel report with color coding
        """
        try:
            # Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Hasil Validasi"

            # Define headers
            headers = [
                'No', 'NIK', 'Nama Anak', 'Tanggal Lahir', 'Bulan', 'Tanggal Ukur',
                'Umur (bulan)', 'Berat (kg)', 'Tinggi (cm)', 'Cara Ukur',
                'Status Berat', 'Status Tinggi', 'Validasi Input', 'Keterangan'
            ]

            # Write headers
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num, value=header)
                cell.font = self.header_font
                cell.fill = self.header_fill
                cell.border = self.thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center')

            # Group measurements by child
            children_data = {}
            for measurement in measurements:
                if measurement.child_id not in children_data:
                    children_data[measurement.child_id] = {
                        'child': measurement.child,
                        'measurements': []
                    }
                children_data[measurement.child_id]['measurements'].append(measurement)

            # Write data
            row_num = 2
            for child_id, data in children_data.items():
                child = data['child']
                child_measurements = sorted(data['measurements'], key=lambda x: x.umur_bulan or 0)

                for i, measurement in enumerate(child_measurements):
                    # Only show child info in first row
                    nik = child.nik if i == 0 else ''
                    nama_anak = child.nama if i == 0 else ''
                    tgl_lahir = child.tgl_lahir.strftime('%d/%m/%Y') if child.tgl_lahir and i == 0 else ''

                    # Format measurement data
                    tgl_ukur = measurement.tgl_ukur.strftime('%d/%m/%Y') if measurement.tgl_ukur else ''
                    berat = f"{measurement.berat:.1f}" if measurement.berat is not None else ''
                    tinggi = f"{measurement.tinggi:.1f}" if measurement.tinggi is not None else ''
                    umur = str(measurement.umur_bulan) if measurement.umur_bulan is not None else ''

                    # Write row data
                    row_data = [
                        row_num - 1,  # No
                        nik,
                        nama_anak,
                        tgl_lahir,
                        measurement.bulan,
                        tgl_ukur,
                        umur,
                        berat,
                        tinggi,
                        measurement.cara_ukur or '',
                        measurement.status_berat,
                        measurement.status_tinggi,
                        measurement.validasi_input,
                        measurement.keterangan or ''
                    ]

                    for col_num, value in enumerate(row_data, 1):
                        cell = ws.cell(row=row_num, column=col_num, value=value)
                        cell.border = self.thin_border

                        # Apply color coding based on validation status
                        if measurement.validasi_input == 'ERROR':
                            cell.fill = self.colors['ERROR']
                        elif measurement.validasi_input == 'WARNING':
                            # Check if it's an anomaly (weight loss > 10%)
                            if 'Anomali berat' in (measurement.keterangan or ''):
                                cell.fill = self.colors['WARNING_ANOMALY']
                            else:
                                cell.fill = self.colors['WARNING']
                        elif measurement.validasi_input == 'OK':
                            cell.fill = self.colors['OK']

                    row_num += 1

            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
                ws.column_dimensions[column_letter].width = adjusted_width

            # Freeze header row
            ws.freeze_panes = 'A2'

            # Add summary sheet
            self._add_summary_sheet(wb, measurements, default_gender)

            # Save workbook
            wb.save(file_path)
            logger.info(f"Excel report saved to {file_path}")

        except Exception as e:
            logger.error(f"Error generating Excel report: {str(e)}")
            raise

    def _add_summary_sheet(self, wb, measurements: List[Measurement], default_gender: str):
        """
        Add summary statistics sheet
        """
        # Create summary sheet
        summary_ws = wb.create_sheet("Ringkasan")

        # Calculate statistics
        total_anak = len(set(m.child_id for m in measurements))
        total_records = len(measurements)
        valid_count = len([m for m in measurements if m.validasi_input == 'OK'])
        warning_count = len([m for m in measurements if m.validasi_input == 'WARNING'])
        error_count = len([m for m in measurements if m.validasi_input == 'ERROR'])
        missing_weight = len([m for m in measurements if m.status_berat == 'Missing'])
        missing_height = len([m for m in measurements if m.status_tinggi == 'Missing'])
        missing_data = missing_weight + missing_height

        # Summary data
        summary_data = [
            ['Parameter', 'Jumlah', 'Persentase'],
            ['Total Anak', total_anak, '100%'],
            ['Total Data Pengukuran', total_records, '100%'],
            ['', '', ''],
            ['Status Validasi', '', ''],
            ['- Valid (OK)', valid_count, f'{valid_count/total_records*100:.1f}%' if total_records > 0 else '0%'],
            ['- Peringatan (Warning)', warning_count, f'{warning_count/total_records*100:.1f}%' if total_records > 0 else '0%'],
            ['- Error', error_count, f'{error_count/total_records*100:.1f}%' if total_records > 0 else '0%'],
            ['', '', ''],
            ['Data Hilang', '', ''],
            ['- Berat Hilang', missing_weight, f'{missing_weight/total_records*100:.1f}%' if total_records > 0 else '0%'],
            ['- Tinggi Hilang', missing_height, f'{missing_height/total_records*100:.1f}%' if total_records > 0 else '0%'],
            ['- Total Data Hilang', missing_data, f'{missing_data/total_records*100:.1f}%' if total_records > 0 else '0%'],
            ['', '', ''],
            ['Informasi Analisis', '', ''],
            ['Jenis Kelamin Default', default_gender, '-'],
            ['Tanggal Generate', datetime.now().strftime('%d/%m/%Y %H:%M:%S'), '-']
        ]

        # Write summary data
        for row_num, row_data in enumerate(summary_data, 1):
            for col_num, value in enumerate(row_data, 1):
                cell = summary_ws.cell(row=row_num, column=col_num, value=value)
                cell.border = self.thin_border

                # Format header row
                if row_num == 1:
                    cell.font = self.header_font
                    cell.fill = self.header_fill
                    cell.alignment = Alignment(horizontal='center', vertical='center')

                # Format category rows
                if value in ['', 'Status Validasi', 'Data Hilang', 'Informasi Analisis'] and col_num == 1:
                    cell.font = Font(bold=True)

        # Auto-adjust column widths
        for column in summary_ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                        pass
            adjusted_width = min(max_length + 2, 50)
            summary_ws.column_dimensions[column_letter].width = adjusted_width

    async def generate_text_report(self, file_path: str, measurements: List[Measurement], summary: Dict):
        """
        Generate descriptive text report per child
        """
        try:
            with open(file_path, 'w', encoding='utf-8') as f:
                # Write header
                f.write("LAPORAN VALIDASI DATA PERTUMBUHAN ANAK\n")
                f.write("=" * 50 + "\n\n")
                f.write(f"Tanggal Generate: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}\n\n")

                # Write summary
                f.write("RINGKASAN ANALISIS\n")
                f.write("-" * 20 + "\n")
                f.write(f"Total Anak: {summary['total_anak']}\n")
                f.write(f"Total Data Pengukuran: {summary['total_records']}\n")
                f.write(f"Valid (OK): {summary['valid']} ({summary['valid']/summary['total_records']*100:.1f}%)\n")
                f.write(f"Peringatan (Warning): {summary['warning']} ({summary['warning']/summary['total_records']*100:.1f}%)\n")
                f.write(f"Error: {summary['error']} ({summary['error']/summary['total_records']*100:.1f}%)\n")
                f.write(f"Missing Data: {summary['missing']}\n\n")

                # Group measurements by child
                children_data = {}
                for measurement in measurements:
                    if measurement.child_id not in children_data:
                        children_data[measurement.child_id] = {
                            'child': measurement.child,
                            'measurements': []
                        }
                    children_data[measurement.child_id]['measurements'].append(measurement)

                # Write detailed report per child
                f.write("ANALISIS DETAIL PER ANAK\n")
                f.write("=" * 30 + "\n\n")

                for child_id, data in children_data.items():
                    child = data['child']
                    child_measurements = sorted(data['measurements'], key=lambda x: x.umur_bulan or 0)

                    # Child information
                    f.write(f"NAMA: {child.nama}\n")
                    if child.nik:
                        f.write(f"NIK: {child.nik}\n")
                    if child.tgl_lahir:
                        f.write(f"TANGGAL LAHIR: {child.tgl_lahir.strftime('%d/%m/%Y')}\n")
                    f.write("-" * 20 + "\n")

                    # Analyze measurements
                    errors = []
                    warnings = []
                    missing_months = []
                    height_issues = []
                    weight_issues = []
                    non_ideal_measurements = []

                    previous_age = None
                    previous_height = None
                    previous_weight = None

                    for measurement in child_measurements:
                        # Check for missing data
                        if measurement.status_berat == 'Missing' or measurement.status_tinggi == 'Missing':
                            missing_months.append(measurement.bulan)

                        # Check for height consistency
                        if measurement.validasi_input == 'ERROR' and 'menurun' in (measurement.keterangan or '').lower():
                            height_issues.append(f"{measurement.keterangan} (Bulan: {measurement.bulan})")

                        # Check for weight anomalies
                        if 'Anomali berat' in (measurement.keterangan or ''):
                            weight_issues.append(f"{measurement.keterangan} (Bulan: {measurement.bulan})")

                        # Check for non-ideal measurements
                        if measurement.status_berat == 'Tidak Ideal' or measurement.status_tinggi == 'Tidak Ideal':
                            non_ideal_measurements.append(measurement)

                        # Check for age gaps
                        if previous_age is not None and measurement.umur_bulan is not None:
                            age_gap = measurement.umur_bulan - previous_age
                            if age_gap > 1:
                                warnings.append(f"Gap data: tidak ada pengukuran untuk {age_gap-1} bulan sebelum {measurement.bulan}")

                        # Update previous values
                        if measurement.umur_bulan is not None:
                            previous_age = measurement.umur_bulan
                        if measurement.tinggi is not None:
                            previous_height = measurement.tinggi
                        if measurement.berat is not None:
                            previous_weight = measurement.berat

                    # Write findings
                    if missing_months:
                        f.write(f"Tidak diukur pada bulan: {', '.join(missing_months)}\n")

                    if height_issues:
                        f.write("\nMASALAH TINGGI BADAN:\n")
                        for issue in height_issues:
                            f.write(f"- {issue}\n")

                    if weight_issues:
                        f.write("\nANOMALI BERAT BADAN:\n")
                        for issue in weight_issues:
                            f.write(f"- {issue}\n")

                    if non_ideal_measurements:
                        f.write("\nDATA DI LUAR RENTANG IDEAL:\n")
                        for m in non_ideal_measurements:
                            if m.status_berat == 'Tidak Ideal':
                                f.write(f"- Berat tidak ideal: {m.berat}kg (Bulan: {m.bulan})\n")
                            if m.status_tinggi == 'Tidak Ideal':
                                f.write(f"- Tinggi tidak ideal: {m.tinggi}cm (Bulan: {m.bulan})\n")

                    if warnings:
                        f.write("\nPERINGATAN:\n")
                        for warning in warnings:
                            f.write(f"- {warning}\n")

                    if not (missing_months or height_issues or weight_issues or non_ideal_measurements or warnings):
                        f.write("SEMUA DATA VALID ✓\n")

                    f.write("\n" + "=" * 50 + "\n\n")

            logger.info(f"Text report saved to {file_path}")

        except Exception as e:
            logger.error(f"Error generating text report: {str(e)}")
            raise